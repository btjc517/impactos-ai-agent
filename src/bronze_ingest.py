"""
Bronze-only ingestion for multi-sheet Excel/CSV with sheet registry.

Deterministic, idempotent, and tenant-scoped.
"""

from __future__ import annotations

import os
import io
import json
import hashlib
from typing import Dict, Any, List, Tuple

import pandas as pd
import sqlite3

from transforms_medallion import ensure_bronze_table, slugify

try:
    from openpyxl import load_workbook  # optional, for hidden detection
    _OPENPYXL_OK = True
except Exception:
    _OPENPYXL_OK = False


def _excel_col_letter(idx_zero_based: int) -> str:
    n = idx_zero_based + 1
    letters = []
    while n:
        n, r = divmod(n - 1, 26)
        letters.append(chr(65 + r))
    return ''.join(reversed(letters))


def _infer_sql_type(dtype: Any) -> str:
    try:
        if pd.api.types.is_integer_dtype(dtype):
            return 'INTEGER'
        if pd.api.types.is_float_dtype(dtype):
            return 'REAL'
        if pd.api.types.is_datetime64_any_dtype(dtype):
            return 'TEXT'  # store ISO strings
        if pd.api.types.is_bool_dtype(dtype):
            return 'INTEGER'
    except Exception:
        pass
    return 'TEXT'


def _hash_dataframe(df: pd.DataFrame) -> str:
    # Normalize values to strings for stable hashing
    normalized = df.copy()
    for col in normalized.columns:
        if pd.api.types.is_datetime64_any_dtype(normalized[col]):
            normalized[col] = normalized[col].dt.strftime('%Y-%m-%dT%H:%M:%S')
        else:
            normalized[col] = normalized[col].astype(str)
    csv_buf = io.StringIO()
    normalized.to_csv(csv_buf, index=False)
    data = ("|".join([str(c) for c in normalized.columns]) + "\n" + csv_buf.getvalue()).encode('utf-8')
    return hashlib.sha256(data).hexdigest()


def _collect_sheet_profile(df: pd.DataFrame) -> Tuple[Dict[str, str], Dict[str, List[str]], List[str], int]:
    column_types: Dict[str, str] = {}
    example_values: Dict[str, List[str]] = {}
    inferred_dates: List[str] = []
    for col in df.columns:
        dtype = df[col].dtype
        sql_type = _infer_sql_type(dtype)
        column_types[str(col)] = sql_type
        # examples: up to 3 non-null stringified values
        ex = [str(v) for v in df[col].dropna().head(3).tolist()]
        example_values[str(col)] = ex
        if pd.api.types.is_datetime64_any_dtype(dtype):
            inferred_dates.append(str(col))
    header_row_idx = 0
    return column_types, example_values, inferred_dates, header_row_idx


def _compute_used_range(df: pd.DataFrame, header_row_idx: int) -> Tuple[int, int, int, int]:
    # Determine used columns (ignore trailing empty columns)
    headers = [str(c).strip() if c is not None else '' for c in df.columns]
    data_nonempty = df.notna().any(axis=0).tolist()
    used = []
    for i, h in enumerate(headers):
        used.append(bool(h) or bool(data_nonempty[i]))
    if not any(used):
        first_col = 0
        last_col = max(0, len(headers) - 1)
    else:
        first_col = next(i for i, v in enumerate(used) if v)
        last_col = len(used) - 1 - next(i for i, v in enumerate(reversed(used)) if v)
    # Determine data rows range within used columns
    if first_col <= last_col:
        used_block = df.iloc[:, first_col:last_col + 1]
        rows_any = used_block.notna().any(axis=1).tolist()
        if any(rows_any):
            first_data_idx = next(i for i, v in enumerate(rows_any) if v)
            last_data_idx = len(rows_any) - 1 - next(i for i, v in enumerate(reversed(rows_any)) if v)
        else:
            first_data_idx = 0
            last_data_idx = max(0, len(rows_any) - 1)
    else:
        first_data_idx = 0
        last_data_idx = 0
    first_row_num = header_row_idx + 1
    first_data_row_num = header_row_idx + 1 + first_data_idx
    last_row_num = header_row_idx + 1 + last_data_idx
    return first_col, last_col, first_data_row_num, last_row_num


def _build_bronze_dataframe(df: pd.DataFrame, source_file: str, sheet_name: str, header_row_idx: int, col_start_letter: str, col_end_letter: str) -> pd.DataFrame:
    df2 = df.copy()
    df2.insert(0, 'tenant_id', 'default')
    df2.insert(1, 'source_file', source_file)
    df2.insert(2, 'sheet_name', sheet_name)
    # row_idx (1-based for data rows, following header)
    df2.insert(3, 'row_idx', df.reset_index(drop=True).index + (header_row_idx + 1))
    # row_cell_ref for each data row using sheet used columns (deterministic)
    ranges: List[str] = []
    for i in range(len(df2)):
        rownum = i + header_row_idx + 1
        ranges.append(f"{col_start_letter}{rownum}:{col_end_letter}{rownum}")
    df2.insert(4, 'row_cell_ref', ranges)
    return df2


def _upsert_sheet_registry(conn: sqlite3.Connection, tenant_id: str, source_file: str, sheet_name: str, bronze_table: str, profile: Tuple[Dict[str, str], Dict[str, List[str]], List[str], int], content_hash: str, data_range: str, header_range: str, has_hidden: bool) -> None:
    col_types, examples, inferred_dates, header_row_idx = profile
    cursor = conn.cursor()
    # Skip if hash exists
    cursor.execute(
        """
        SELECT id FROM sheet_registry
        WHERE tenant_id = ? AND source_file = ? AND sheet_name = ? AND content_hash = ?
        """,
        (tenant_id, source_file, sheet_name, content_hash),
    )
    if cursor.fetchone():
        return
    # Determine next version
    cursor.execute(
        """
        SELECT COALESCE(MAX(version), 0) FROM sheet_registry
        WHERE tenant_id = ? AND source_file = ? AND sheet_name = ?
        """,
        (tenant_id, source_file, sheet_name),
    )
    next_version = int(cursor.fetchone()[0]) + 1
    cursor.execute(
        """
        INSERT INTO sheet_registry
        (tenant_id, source_file, sheet_name, bronze_table, column_types, example_values, inferred_date_cols, header_row_idx, content_hash, version, data_range, header_range, has_hidden)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            tenant_id,
            source_file,
            sheet_name,
            bronze_table,
            json.dumps(col_types),
            json.dumps(examples),
            json.dumps(inferred_dates),
            header_row_idx,
            content_hash,
            next_version,
            data_range,
            header_range,
            1 if has_hidden else 0,
        ),
    )
    conn.commit()


def ingest_bronze(file_path: str, db_path: str = 'db/impactos.db', tenant_id: str = 'default') -> Dict[str, Any]:
    # Default to extension-based detection
    is_excel = str(file_path).lower().endswith(('.xlsx', '.xlsm', '.xls'))
    is_csv = str(file_path).lower().endswith('.csv')
    # Try semantic resolver; if it accepts, override; if it abstains, keep defaults
    try:
        from semantic_resolver import SemanticResolver
        r = SemanticResolver().resolve_file_type(str(file_path))
        if r.get('outcome') == 'accepted':
            key = r.get('key')
            is_excel = (key == 'excel')
            is_csv = (key == 'csv')
    except Exception:
        pass
    if not (is_excel or is_csv):
        raise ValueError('Only Excel or CSV supported for Bronze ingestion')
    source_file = os.path.basename(file_path)
    with sqlite3.connect(db_path) as conn:
        conn.execute("PRAGMA foreign_keys = ON")
        # Ensure required DDL applied (core + sheet registry + transforms), idempotently
        proj_root = os.path.dirname(os.path.dirname(__file__))
        # medallion core
        mig_core = os.path.join(proj_root, 'db', 'migrations', '20250813T000000Z__medallion_v1.sql')
        if os.path.exists(mig_core):
            with open(mig_core, 'r') as f:
                conn.executescript(f.read())
        # transform support migrations
        for mig in [
            '20250813T040000Z__transform_runs.sql',
            '20250813T041000Z__fact_meta_and_dim_uniques.sql',
            '20250813T042000Z__transform_rejects_lineage.sql',
            '20250813T050000Z__transform_jobs.sql',
        ]:
            fp = os.path.join(proj_root, 'db', 'migrations', mig)
            if os.path.exists(fp):
                with open(fp, 'r') as f:
                    try:
                        conn.executescript(f.read())
                    except Exception:
                        pass
        # sheet_registry base if table missing
        cur = conn.cursor()
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='sheet_registry'")
        has_registry = cur.fetchone() is not None
        if not has_registry:
            mig_reg = os.path.join(proj_root, 'db', 'migrations', '20250813T020000Z__sheet_registry.sql')
            if os.path.exists(mig_reg):
                with open(mig_reg, 'r') as f:
                    conn.executescript(f.read())
        # add ranges columns only if missing
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='sheet_registry'")
        if cur.fetchone() is not None:
            cur.execute("PRAGMA table_info(sheet_registry)")
            cols = {r[1] for r in cur.fetchall()}
            if 'data_range' not in cols or 'header_range' not in cols or 'has_hidden' not in cols:
                mig_rng = os.path.join(proj_root, 'db', 'migrations', '20250813T030000Z__sheet_registry_ranges.sql')
                if os.path.exists(mig_rng):
                    with open(mig_rng, 'r') as f:
                        try:
                            conn.executescript(f.read())
                        except Exception:
                            pass

        created_tables: List[str] = []
        registry_rows: List[Dict[str, Any]] = []

        if is_excel:
            sheets: Dict[str, pd.DataFrame] = pd.read_excel(file_path, sheet_name=None, header=0)
        else:
            # Single pseudo-sheet for CSV
            df = pd.read_csv(file_path)
            sheets = {'Sheet1': df}

        for sheet_name, df in sheets.items():
            # Profile + hash
            profile = _collect_sheet_profile(df)
            content_hash = _hash_dataframe(df)
            # Used range computation
            first_col_idx, last_col_idx, first_data_row_num, last_row_num = _compute_used_range(df, profile[3])
            col_start_letter = _excel_col_letter(first_col_idx)
            col_end_letter = _excel_col_letter(last_col_idx)
            header_row_num = profile[3] + 1
            data_range = f"{col_start_letter}{header_row_num}:{col_end_letter}{last_row_num}"
            header_range = f"{col_start_letter}{header_row_num}:{col_end_letter}{header_row_num}"
            # Hidden detection (Excel only)
            has_hidden = False
            if is_excel and _OPENPYXL_OK:
                try:
                    wb = load_workbook(file_path, read_only=True, data_only=True)
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        # rows
                        for r in range(header_row_num, last_row_num + 1):
                            rd = ws.row_dimensions.get(r)
                            if rd and getattr(rd, 'hidden', False):
                                has_hidden = True
                                break
                        # columns
                        if not has_hidden:
                            for c in range(first_col_idx, last_col_idx + 1):
                                cd = ws.column_dimensions.get(_excel_col_letter(c))
                                if cd and getattr(cd, 'hidden', False):
                                    has_hidden = True
                                    break
                except Exception:
                    has_hidden = False

            # Prepare bronze table and DataFrame
            file_slug = slugify(os.path.splitext(source_file)[0])
            sheet_slug = slugify(sheet_name)
            bronze_table = f"bronze_{file_slug}__{sheet_slug}"

            # Build slugified, de-duplicated column map for Bronze schema
            rename_map: Dict[str, str] = {}
            seen: Dict[str, int] = {}
            for col in df.columns:
                base = slugify(str(col))
                name = base
                # Ensure uniqueness if multiple columns slugify to the same token
                if name in seen:
                    seen[name] += 1
                    name = f"{base}_{seen[base]}"
                else:
                    seen[name] = 1
                rename_map[str(col)] = name
            raw_types = {rename_map[str(c)]: _infer_sql_type(df[c].dtype) for c in df.columns}
            ensure_bronze_table(conn, file_slug=file_slug, sheet_slug=sheet_slug, raw_columns=raw_types)

            bronze_df = _build_bronze_dataframe(
                df,
                source_file,
                sheet_name,
                header_row_idx=profile[3],
                col_start_letter=col_start_letter,
                col_end_letter=col_end_letter,
            )
            # Align DataFrame column names with Bronze table schema (slugified, de-duplicated)
            try:
                if rename_map:
                    bronze_df.rename(columns=rename_map, inplace=True)
            except Exception:
                pass

            # Idempotency: skip if hash exists for this sheet
            cur = conn.cursor()
            cur.execute(
                "SELECT id FROM sheet_registry WHERE tenant_id=? AND source_file=? AND sheet_name=? AND content_hash=?",
                (tenant_id, source_file, sheet_name, content_hash),
            )
            row_existing = cur.fetchone()
            exists = row_existing is not None
            if not exists:
                # Append bronze rows
                # Ensure all columns present as strings where necessary
                # Convert datetimes to ISO strings to preserve raw values
                for col in bronze_df.columns:
                    if pd.api.types.is_datetime64_any_dtype(bronze_df[col]):
                        bronze_df[col] = bronze_df[col].dt.strftime('%Y-%m-%dT%H:%M:%S')
                bronze_df.to_sql(bronze_table, conn, if_exists='append', index=False)
                _upsert_sheet_registry(conn, tenant_id, source_file, sheet_name, bronze_table, profile, content_hash, data_range, header_range, has_hidden)
                conn.commit()

            # Auto-enqueue and optionally run transforms (also on re-ingest of same content)
            auto = (os.getenv('AUTO_TRANSFORM', 'true').lower() in ('1','true','yes'))
            mode = os.getenv('TRANSFORM_MODE', 'sync').lower()
            if auto:
                # lookup sheet_registry id (existing or newly inserted)
                cur2 = conn.cursor()
                cur2.execute("SELECT id FROM sheet_registry WHERE tenant_id=? AND source_file=? AND sheet_name=? AND content_hash=? ORDER BY id DESC LIMIT 1",
                             (tenant_id, source_file, sheet_name, content_hash))
                rowid = cur2.fetchone()
                sheet_id = rowid[0] if rowid else None
                
                # FLEXIBLE PIPELINE PROCESSING: Use dynamic pipeline instead of hardcoded transforms
                use_flexible_pipeline = os.getenv('ENABLE_FLEXIBLE_PIPELINE', 'true').lower() in ('1','true','yes')
                if use_flexible_pipeline:
                    try:
                        from flexible_pipeline import FlexiblePipelineEngine
                        
                        # Initialize pipeline engine
                        pipeline_engine = FlexiblePipelineEngine(db_path=db_path)
                        
                        # Determine which pipeline to use
                        pipeline_id = os.getenv('PROCESSING_PIPELINE', 'dynamic_processing')
                        
                        # Prepare initial data for pipeline
                        initial_data = {
                            'dataframe': df,
                            'source_file': source_file,
                            'sheet_name': sheet_name,
                            'tenant_id': tenant_id,
                            'bronze_table': bronze_table,
                            'sheet_id': sheet_id,
                            'content_hash': content_hash,
                            'is_new_data': not exists
                        }
                        
                        # Execute flexible pipeline
                        run_id = pipeline_engine.execute_pipeline(pipeline_id, initial_data)
                        print(f"🚀 Executed flexible pipeline '{pipeline_id}' for {sheet_name} (run: {run_id})")
                        
                        # Skip traditional processing since pipeline handles it
                        continue
                        
                    except Exception as e:
                        print(f"Warning: Flexible pipeline failed for {sheet_name}, falling back to traditional processing: {e}")
                
                # FALLBACK: Traditional processing with dynamic fact discovery
                dynamic_fact_discovery = os.getenv('ENABLE_FACT_DISCOVERY', 'true').lower() in ('1','true','yes')
                if dynamic_fact_discovery and not exists:  # Only discover on new data
                    try:
                        from spec_resolver import FactsConfig
                        facts_config = FactsConfig(db_path=db_path)
                        
                        # Create source context for fact discovery
                        source_context = {
                            'source_file': source_file,
                            'sheet_name': sheet_name,
                            'tenant_id': tenant_id,
                            'columns': list(df.columns),
                            'row_count': len(df)
                        }
                        
                        # Discover new facts from this sheet's data
                        discovered_count = facts_config.discover_facts_from_data(df, source_context)
                        
                        if discovered_count > 0:
                            print(f"🧠 Discovered {discovered_count} new fact patterns from {sheet_name}")
                        
                    except Exception as e:
                        print(f"Warning: Fact discovery failed for {sheet_name}: {e}")
                
                # Resolve specs (traditional approach)
                from spec_resolver import resolve_specs_for_sheet
                # Provide bronze-aligned headers (slugified) to resolver so mappings match table columns
                sheet_headers = [slugify(str(c)) for c in df.columns]
                
                # Build example values for spec resolver
                example_values = {}
                for col in df.columns:
                    original_col = str(col)
                    slugified_col = slugify(original_col)
                    examples = df[col].dropna().astype(str).head(3).tolist()
                    example_values[slugified_col] = examples
                
                specs = resolve_specs_for_sheet(tenant_id, {
                    'sheet_name': sheet_name,
                    'bronze_table': bronze_table,
                    'source_file': source_file,
                    'columns': sheet_headers,
                    'example_values': example_values,
                })
                enqueued = 0
                for s in specs:
                    try:
                        # Skip enqueue if a succeeded job with same spec already exists for this sheet
                        cur2.execute(
                            "SELECT id, status FROM transform_jobs WHERE tenant_id=? AND sheet_registry_id=? AND spec_hash=? ORDER BY id DESC LIMIT 1",
                            (tenant_id, sheet_id, s['spec_hash'])
                        )
                        prior = cur2.fetchone()
                        if prior and (prior[1] or '').lower() == 'succeeded' and exists:
                            continue
                        conn.execute(
                            """
                            INSERT INTO transform_jobs (tenant_id, sheet_registry_id, bronze_table, spec_id, spec_hash)
                            VALUES (?,?,?,?,?)
                            """,
                            (tenant_id, sheet_id, bronze_table, s['spec_id'], s['spec_hash'])
                        )
                        enqueued += 1
                        if mode == 'sync':
                            # run immediately
                            conn.execute("UPDATE transform_jobs SET status='running', started_at=CURRENT_TIMESTAMP WHERE tenant_id=? AND sheet_registry_id=? AND spec_hash=?", (tenant_id, sheet_id, s['spec_hash']))
                            from silver_transform import run_transform
                            spec_snapshot = s['spec_snapshot']
                            res = run_transform(conn, bronze_table, spec_snapshot, tenant_id=tenant_id)
                            # Fallback: try title-cased headers if no rows out
                            if not res.get('rows_out'):
                                snap2 = json.loads(json.dumps(spec_snapshot))
                                mp = snap2.get('mappings', {})
                                for k, v in list(mp.items()):
                                    if isinstance(v, str):
                                        mp[k] = v.title()
                                res = run_transform(conn, bronze_table, snap2, tenant_id=tenant_id)
                            # link run to job + sheet
                            conn.execute(
                                "UPDATE transform_runs SET transform_job_id=(SELECT id FROM transform_jobs WHERE tenant_id=? AND sheet_registry_id=? AND spec_hash=?), sheet_registry_id=? WHERE id=?",
                                (tenant_id, sheet_id, s['spec_hash'], sheet_id, res.get('run_id'))
                            )
                            conn.execute("UPDATE transform_jobs SET status='succeeded', finished_at=CURRENT_TIMESTAMP WHERE tenant_id=? AND sheet_registry_id=? AND spec_hash=?",
                                         (tenant_id, sheet_id, s['spec_hash']))
                    except Exception as e:
                        # duplicate enqueue or failure
                        try:
                            conn.execute("UPDATE transform_jobs SET status='failed', attempts=attempts+1, last_error=? WHERE tenant_id=? AND sheet_registry_id=? AND spec_hash=?",
                                         (str(e)[:400], tenant_id, sheet_id, s['spec_hash']))
                        except Exception:
                            pass
                # Log count (stdout)
                print(f"Auto-transform: enqueued={enqueued} mode={mode} for sheet_id={sheet_id}")

            created_tables.append(bronze_table)
            registry_rows.append({
                'tenant_id': tenant_id,
                'source_file': source_file,
                'sheet_name': sheet_name,
                'bronze_table': bronze_table,
                'hash': content_hash,
                'data_range': data_range,
                'header_range': header_range,
                'has_hidden': has_hidden,
            })

        return {
            'created_tables': list(sorted(set(created_tables))),
            'registry_rows': registry_rows,
        }


