"""
Enhanced data loader for bulletproof ingestion accuracy.

This module implements advanced data extraction strategies:
- Pre-scan columns with openpyxl for type inference
- Handle mixed types via custom converters
- Use Polars for type safety with schema overrides
- Capture precise cell-level references
"""

import pandas as pd
import polars as pl
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from typing import Dict, List, Any, Optional, Tuple, Union
from pathlib import Path
import logging
from datetime import datetime, date
import json
import re
from decimal import Decimal, InvalidOperation

logger = logging.getLogger(__name__)


class CellReference:
    """Represents a precise cell reference with Excel coordinates."""
    
    def __init__(self, row: int, col: int, sheet_name: str = "Sheet1", 
                 value: Any = None, formatted_value: str = None, 
                 data_type: str = None):
        self.row = row  # 1-based Excel row
        self.col = col  # 1-based Excel column
        self.sheet_name = sheet_name
        self.value = value
        self.formatted_value = formatted_value
        self.data_type = data_type
        
    @property
    def excel_address(self) -> str:
        """Get Excel-style address like 'A1', 'B5', etc."""
        col_letter = get_column_letter(self.col)
        return f"{col_letter}{self.row}"
    
    @property
    def address_with_sheet(self) -> str:
        """Get full address with sheet name like 'Sheet1!A1'."""
        return f"{self.sheet_name}!{self.excel_address}"
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary for storage."""
        return {
            'row': self.row,
            'col': self.col,
            'sheet_name': self.sheet_name,
            'excel_address': self.excel_address,
            'address_with_sheet': self.address_with_sheet,
            'value': self.value,
            'formatted_value': self.formatted_value,
            'data_type': self.data_type
        }


class ColumnTypeInference:
    """Handles intelligent type inference for Excel columns."""
    
    @staticmethod
    def infer_column_type(cells: List[openpyxl.cell.Cell]) -> Dict[str, Any]:
        """
        Analyze a column of cells to infer the most appropriate data type.
        
        Returns:
            Dict with type info, converter function, and metadata
        """
        non_empty_cells = [cell for cell in cells if cell.value is not None]
        
        if not non_empty_cells:
            return {
                'inferred_type': 'string',
                'converter': str,
                'confidence': 0.0,
                'sample_values': [],
                'issues': ['All cells are empty']
            }
        
        # Collect sample values and their types
        sample_values = [cell.value for cell in non_empty_cells[:10]]
        cell_types = [type(cell.value).__name__ for cell in non_empty_cells]
        
        # Count occurrences of each type
        type_counts = {}
        for cell_type in cell_types:
            type_counts[cell_type] = type_counts.get(cell_type, 0) + 1
        
        total_cells = len(non_empty_cells)
        issues = []
        
        # Determine predominant type and confidence
        if len(type_counts) == 1:
            # All cells have the same type
            predominant_type = list(type_counts.keys())[0]
            confidence = 1.0
        else:
            # Mixed types - choose most common
            predominant_type = max(type_counts, key=type_counts.get)
            confidence = type_counts[predominant_type] / total_cells
            issues.append(f"Mixed types detected: {type_counts}")
        
        # Map Python types to pandas/polars types with converters
        type_mapping = {
            'int': {
                'inferred_type': 'int64',
                'converter': ColumnTypeInference._safe_int_converter,
                'polars_type': pl.Int64
            },
            'float': {
                'inferred_type': 'float64',
                'converter': ColumnTypeInference._safe_float_converter,
                'polars_type': pl.Float64
            },
            'str': {
                'inferred_type': 'string',
                'converter': ColumnTypeInference._safe_string_converter,
                'polars_type': pl.Utf8
            },
            'datetime': {
                'inferred_type': 'datetime64[ns]',
                'converter': ColumnTypeInference._safe_datetime_converter,
                'polars_type': pl.Datetime
            },
            'date': {
                'inferred_type': 'datetime64[ns]',
                'converter': ColumnTypeInference._safe_datetime_converter,
                'polars_type': pl.Date
            },
            'bool': {
                'inferred_type': 'boolean',
                'converter': ColumnTypeInference._safe_bool_converter,
                'polars_type': pl.Boolean
            }
        }
        
        # Handle special cases for numeric strings
        if predominant_type == 'str':
            numeric_pattern = re.compile(r'^-?\d+\.?\d*$')
            currency_pattern = re.compile(r'^[\$£€¥]?-?\d{1,3}(,\d{3})*\.?\d*$')
            
            numeric_count = sum(1 for cell in non_empty_cells 
                              if isinstance(cell.value, str) and numeric_pattern.match(cell.value.strip()))
            currency_count = sum(1 for cell in non_empty_cells 
                               if isinstance(cell.value, str) and currency_pattern.match(cell.value.strip()))
            
            if numeric_count / total_cells > 0.8:
                issues.append("String column contains mostly numeric values")
                return {
                    'inferred_type': 'float64',
                    'converter': ColumnTypeInference._numeric_string_converter,
                    'polars_type': pl.Float64,
                    'confidence': confidence * 0.9,
                    'sample_values': sample_values,
                    'issues': issues
                }
            elif currency_count / total_cells > 0.8:
                issues.append("String column contains mostly currency values")
                return {
                    'inferred_type': 'float64',
                    'converter': ColumnTypeInference._currency_string_converter,
                    'polars_type': pl.Float64,
                    'confidence': confidence * 0.9,
                    'sample_values': sample_values,
                    'issues': issues
                }
        
        # Default mapping
        type_info = type_mapping.get(predominant_type, type_mapping['str'])
        
        return {
            **type_info,
            'confidence': confidence,
            'sample_values': sample_values,
            'issues': issues,
            'type_distribution': type_counts
        }
    
    @staticmethod
    def _safe_int_converter(value):
        """Safely convert value to integer."""
        if pd.isna(value) or value is None:
            return None
        try:
            if isinstance(value, (int, float)):
                return int(value)
            elif isinstance(value, str):
                cleaned = value.strip().replace(',', '')
                return int(float(cleaned))
            return int(value)
        except (ValueError, TypeError):
            return None
    
    @staticmethod
    def _safe_float_converter(value):
        """Safely convert value to float."""
        if pd.isna(value) or value is None:
            return None
        try:
            if isinstance(value, (int, float)):
                return float(value)
            elif isinstance(value, str):
                cleaned = value.strip().replace(',', '')
                return float(cleaned)
            return float(value)
        except (ValueError, TypeError):
            return None
    
    @staticmethod
    def _safe_string_converter(value):
        """Safely convert value to string."""
        if pd.isna(value) or value is None:
            return None
        return str(value).strip()
    
    @staticmethod
    def _safe_datetime_converter(value):
        """Safely convert value to datetime."""
        if pd.isna(value) or value is None:
            return None
        try:
            if isinstance(value, (datetime, date)):
                return pd.to_datetime(value)
            elif isinstance(value, str):
                return pd.to_datetime(value, infer_datetime_format=True)
            return pd.to_datetime(value)
        except (ValueError, TypeError):
            return None
    
    @staticmethod
    def _safe_bool_converter(value):
        """Safely convert value to boolean."""
        if pd.isna(value) or value is None:
            return None
        try:
            if isinstance(value, bool):
                return value
            elif isinstance(value, str):
                lower_val = value.lower().strip()
                if lower_val in ['true', 'yes', '1', 'on', 'y']:
                    return True
                elif lower_val in ['false', 'no', '0', 'off', 'n']:
                    return False
            elif isinstance(value, (int, float)):
                return bool(value)
            return bool(value)
        except (ValueError, TypeError):
            return None
    
    @staticmethod
    def _numeric_string_converter(value):
        """Convert numeric strings to float."""
        if pd.isna(value) or value is None:
            return None
        try:
            if isinstance(value, (int, float)):
                return float(value)
            elif isinstance(value, str):
                cleaned = value.strip().replace(',', '')
                return float(cleaned)
            return float(value)
        except (ValueError, TypeError):
            return None
    
    @staticmethod
    def _currency_string_converter(value):
        """Convert currency strings to float."""
        if pd.isna(value) or value is None:
            return None
        try:
            if isinstance(value, (int, float)):
                return float(value)
            elif isinstance(value, str):
                # Remove currency symbols and commas
                cleaned = re.sub(r'[\$£€¥,]', '', value.strip())
                return float(cleaned)
            return float(value)
        except (ValueError, TypeError):
            return None


class EnhancedFileLoader:
    """
    Enhanced file loader with bulletproof data extraction strategies.
    
    Features:
    - Pre-scan columns with openpyxl for type inference
    - Handle mixed types via custom converters
    - Use Polars for type safety with schema overrides
    - Capture precise cell-level references
    """
    
    def __init__(self, use_polars: bool = True):
        """
        Initialize enhanced loader.
        
        Args:
            use_polars: Whether to use Polars for type safety (default: True)
        """
        self.use_polars = use_polars
        self.cell_references = {}  # Store cell reference mappings
        self.type_inference_results = {}
        self.load_metadata = {}
        
    def load_file(self, file_path: str, sheet_name: Optional[str] = None) -> Tuple[Union[pd.DataFrame, pl.DataFrame], Dict[str, Any]]:
        """
        Load file with enhanced accuracy and cell reference tracking.
        
        Args:
            file_path: Path to the file to load
            sheet_name: Specific sheet to load (for Excel files)
            
        Returns:
            Tuple of (DataFrame, metadata dict with cell references and type info)
        """
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        file_ext = file_path.suffix.lower()
        # Allow semantic resolution for file type to avoid brittle checks
        try:
            from semantic_resolver import SemanticResolver
            res = SemanticResolver().resolve_file_type(str(file_path))
            if res.get('outcome') == 'accepted':
                key = res.get('key')
                if key == 'excel':
                    return self._load_excel_enhanced(file_path, sheet_name)
                if key == 'csv':
                    return self._load_csv_enhanced(file_path)
        except Exception:
            pass
        # Fallback to extension-based
        if file_ext == '.xlsx':
            return self._load_excel_enhanced(file_path, sheet_name)
        if file_ext == '.csv':
            return self._load_csv_enhanced(file_path)
        raise ValueError(f"Unsupported file format: {file_ext}")
    
    def _load_excel_enhanced(self, file_path: Path, sheet_name: Optional[str] = None) -> Tuple[Union[pd.DataFrame, pl.DataFrame], Dict[str, Any]]:
        """Load Excel file with openpyxl pre-scanning and type inference."""
        logger.info(f"Loading Excel file with enhanced accuracy: {file_path}")
        
        # Step 1: Pre-scan with openpyxl
        workbook = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
        
        # Determine which sheet to use
        if sheet_name is None:
            sheet_name = workbook.active.title
        
        if sheet_name not in workbook.sheetnames:
            available_sheets = ", ".join(workbook.sheetnames)
            raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {available_sheets}")
        
        worksheet = workbook[sheet_name]
        
        # Step 2: Analyze structure and gather cell references (with empty-column trimming)
        cell_data, structure_info = self._analyze_worksheet_structure(worksheet, sheet_name)
        
        # Step 3: Perform type inference on columns
        type_inference = self._perform_type_inference(cell_data, structure_info)
        
        # Step 4: Create custom converters for pandas
        converters = self._create_pandas_converters(type_inference)
        
        # Step 5: Load with pandas using custom converters
        try:
            df_pandas = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                converters=converters,
                na_values=['', 'NULL', 'null', 'N/A', 'n/a', '#N/A', 'NA'],
                keep_default_na=True
            )
            
            # Step 6: Optionally convert to Polars with schema enforcement
            if self.use_polars:
                df_result = self._convert_to_polars_safe(df_pandas, type_inference)
            else:
                df_result = df_pandas
            
            # Step 7: Prepare metadata
            metadata = {
                'file_path': str(file_path),
                'sheet_name': sheet_name,
                'structure_info': structure_info,
                'type_inference': type_inference,
                'cell_references': self.cell_references[sheet_name],
                'load_method': 'polars' if self.use_polars else 'pandas',
                'total_rows': len(df_result),
                'total_columns': len(df_result.columns),
                'load_timestamp': datetime.now().isoformat()
            }
            
            logger.info(f"Successfully loaded {len(df_result)} rows, {len(df_result.columns)} columns with enhanced accuracy")
            
            return df_result, metadata
            
        except Exception as e:
            logger.error(f"Error loading Excel file: {e}")
            raise
        finally:
            workbook.close()
    
    def _load_csv_enhanced(self, file_path: Path) -> Tuple[Union[pd.DataFrame, pl.DataFrame], Dict[str, Any]]:
        """Load CSV file with type inference and safe conversion."""
        logger.info(f"Loading CSV file with enhanced accuracy: {file_path}")
        
        # Step 1: Sample the CSV to infer types
        sample_df = pd.read_csv(file_path, nrows=100)
        
        # Step 2: Analyze column types from sample
        type_inference = {}
        for col in sample_df.columns:
            # Convert to openpyxl-like cell objects for consistency
            cells = [type('MockCell', (), {'value': val})() for val in sample_df[col].dropna()]
            type_inference[col] = ColumnTypeInference.infer_column_type(cells)
        
        # Step 3: Create converters
        converters = self._create_pandas_converters(type_inference)
        
        # Step 4: Load full CSV with converters
        try:
            df_pandas = pd.read_csv(
                file_path,
                converters=converters,
                na_values=['', 'NULL', 'null', 'N/A', 'n/a', '#N/A', 'NA'],
                keep_default_na=True
            )
            
            # Step 5: Optionally convert to Polars
            if self.use_polars:
                df_result = self._convert_to_polars_safe(df_pandas, type_inference)
            else:
                df_result = df_pandas
            
            # Step 6: Create basic cell references for CSV
            self._create_csv_cell_references(df_pandas, file_path.stem)
            
            # Step 7: Prepare metadata
            metadata = {
                'file_path': str(file_path),
                'sheet_name': file_path.stem,
                'type_inference': type_inference,
                'cell_references': self.cell_references.get(file_path.stem, {}),
                'load_method': 'polars' if self.use_polars else 'pandas',
                'total_rows': len(df_result),
                'total_columns': len(df_result.columns),
                'load_timestamp': datetime.now().isoformat()
            }
            
            logger.info(f"Successfully loaded CSV with {len(df_result)} rows, {len(df_result.columns)} columns")
            
            return df_result, metadata
            
        except Exception as e:
            logger.error(f"Error loading CSV file: {e}")
            raise
    
    def _analyze_worksheet_structure(self, worksheet, sheet_name: str) -> Tuple[Dict[str, List], Dict[str, Any]]:
        """Analyze worksheet structure and extract cell data with references.
        
        Trims trailing all-empty columns and excludes any all-empty columns from
        headers, type inference, and metadata to prevent phantom columns.
        """
        # Find the raw sheet bounds
        min_row = worksheet.min_row
        max_row = worksheet.max_row
        min_col = worksheet.min_column
        max_col = worksheet.max_column

        # Identify columns that contain any non-empty cell in data rows
        def _is_non_empty(v: Any) -> bool:
            return v is not None and (not isinstance(v, str) or v.strip() != "")

        has_data_by_col: Dict[int, bool] = {}
        for col in range(min_col, max_col + 1):
            found = False
            for row in range(min_row + 1, max_row + 1):
                v = worksheet.cell(row=row, column=col).value
                if _is_non_empty(v):
                    found = True
                    break
            has_data_by_col[col] = found

        # Determine rightmost column with data, then build include list
        cols_with_data = [c for c in range(min_col, max_col + 1) if has_data_by_col.get(c, False)]
        if cols_with_data:
            rightmost_with_data = max(cols_with_data)
            columns_to_consider = range(min_col, rightmost_with_data + 1)
        else:
            # No data at all: consider zero columns
            columns_to_consider = []

        # Final columns to include: only those with data within the considered range
        include_columns = [c for c in columns_to_consider if has_data_by_col.get(c, False)]

        # Extract all cell data with references for included columns only
        cell_data: Dict[str, List] = {}
        cell_references: Dict[str, List[CellReference]] = {}
        headers: List[str] = []

        for col in include_columns:
            cell_header = worksheet.cell(row=min_row, column=col)
            header_value = cell_header.value if _is_non_empty(cell_header.value) else f"Column_{col}"
            header_value = str(header_value)
            headers.append(header_value)
            cell_data[header_value] = []
            cell_references[header_value] = []

        for row in range(min_row + 1, max_row + 1):
            for col in include_columns:
                header = headers[include_columns.index(col)]
                cell = worksheet.cell(row=row, column=col)
                cell_ref = CellReference(
                    row=row,
                    col=col,
                    sheet_name=sheet_name,
                    value=cell.value,
                    formatted_value=str(cell.value) if cell.value is not None else None,
                    data_type=type(cell.value).__name__ if cell.value is not None else 'NoneType'
                )
                cell_data[header].append(cell)
                cell_references[header].append(cell_ref)

        # Store cell references for later use
        self.cell_references[sheet_name] = cell_references

        total_columns = len(include_columns)
        sheet_dim_right = include_columns[-1] if include_columns else min_col
        structure_info = {
            'total_rows': max_row - min_row,
            'total_columns': total_columns,
            'data_start_row': min_row + 1,
            'data_end_row': max_row,
            'headers': headers,
            'sheet_dimensions': f"{get_column_letter(min_col)}{min_row}:{get_column_letter(sheet_dim_right)}{max_row}" if include_columns else f"{get_column_letter(min_col)}{min_row}:{get_column_letter(min_col)}{min_row}"
        }

        return cell_data, structure_info
    
    def _perform_type_inference(self, cell_data: Dict[str, List], structure_info: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
        """Perform intelligent type inference on all columns."""
        
        type_inference = {}
        
        for column_name, cells in cell_data.items():
            logger.debug(f"Analyzing column: {column_name}")
            
            # Perform type inference
            inference_result = ColumnTypeInference.infer_column_type(cells)
            type_inference[column_name] = inference_result
            
            # Log insights
            issues = inference_result.get('issues', []) or []
            if any('All cells are empty' in str(issue) for issue in issues):
                logger.debug(f"Empty column excluded or detected for '{column_name}'")
            elif inference_result['confidence'] < 0.8:
                logger.warning(f"Low confidence ({inference_result['confidence']:.2f}) for column '{column_name}': {issues}")
            
            logger.debug(f"Column '{column_name}' inferred as {inference_result['inferred_type']} (confidence: {inference_result['confidence']:.2f})")
        
        self.type_inference_results = type_inference
        return type_inference
    
    def _create_pandas_converters(self, type_inference: Dict[str, Dict[str, Any]]) -> Dict[str, callable]:
        """Create custom converters for pandas based on type inference."""
        
        converters = {}
        
        for column_name, inference in type_inference.items():
            if inference['confidence'] > 0.5:  # Only use converters for confident inferences
                converters[column_name] = inference['converter']
                logger.debug(f"Created converter for column '{column_name}': {inference['inferred_type']}")
        
        return converters
    
    def _convert_to_polars_safe(self, df_pandas: pd.DataFrame, type_inference: Dict[str, Dict[str, Any]]) -> pl.DataFrame:
        """Safely convert pandas DataFrame to Polars with schema overrides."""
        
        try:
            # Create Polars schema based on type inference
            schema_overrides = {}
            
            for column_name, inference in type_inference.items():
                if column_name in df_pandas.columns and 'polars_type' in inference:
                    schema_overrides[column_name] = inference['polars_type']
            
            # Convert to Polars with schema enforcement
            df_polars = pl.from_pandas(df_pandas, schema_overrides=schema_overrides)
            
            logger.info(f"Successfully converted to Polars with schema overrides for {len(schema_overrides)} columns")
            
            return df_polars
            
        except Exception as e:
            logger.warning(f"Failed to convert to Polars with schema overrides: {e}. Using default conversion.")
            
            # Fallback to basic conversion
            try:
                return pl.from_pandas(df_pandas)
            except Exception as e2:
                logger.error(f"Failed to convert to Polars: {e2}. Returning pandas DataFrame.")
                return df_pandas
    
    def _create_csv_cell_references(self, df: pd.DataFrame, sheet_name: str):
        """Create cell references for CSV data (simulating Excel-like references)."""
        
        cell_references = {}
        
        for col_idx, column_name in enumerate(df.columns):
            cell_refs = []
            
            for row_idx in range(len(df)):
                # Create cell reference (1-based indexing like Excel)
                cell_ref = CellReference(
                    row=row_idx + 2,  # +2 because row 1 is header, and Excel is 1-based
                    col=col_idx + 1,  # +1 because Excel is 1-based
                    sheet_name=sheet_name,
                    value=df.iloc[row_idx, col_idx],
                    formatted_value=str(df.iloc[row_idx, col_idx]) if pd.notna(df.iloc[row_idx, col_idx]) else None,
                    data_type=type(df.iloc[row_idx, col_idx]).__name__
                )
                
                cell_refs.append(cell_ref)
            
            cell_references[column_name] = cell_refs
        
        self.cell_references[sheet_name] = cell_references
    
    def get_cell_reference(self, sheet_name: str, column_name: str, row_index: int) -> Optional[CellReference]:
        """Get specific cell reference by sheet, column, and row index."""
        
        if sheet_name not in self.cell_references:
            return None
        
        if column_name not in self.cell_references[sheet_name]:
            return None
        
        cell_refs = self.cell_references[sheet_name][column_name]
        
        if 0 <= row_index < len(cell_refs):
            return cell_refs[row_index]
        
        return None
    
    def find_cells_by_value(self, sheet_name: str, value: Any, column_name: Optional[str] = None) -> List[CellReference]:
        """Find all cells containing a specific value."""
        
        matches = []
        
        if sheet_name not in self.cell_references:
            return matches
        
        columns_to_search = [column_name] if column_name else self.cell_references[sheet_name].keys()
        
        for col in columns_to_search:
            if col in self.cell_references[sheet_name]:
                for cell_ref in self.cell_references[sheet_name][col]:
                    if cell_ref.value == value:
                        matches.append(cell_ref)
        
        return matches
    
    def export_cell_references_json(self, output_path: str):
        """Export all cell references to JSON for debugging/verification."""
        
        export_data = {}
        
        for sheet_name, sheet_refs in self.cell_references.items():
            export_data[sheet_name] = {}
            
            for column_name, cell_refs in sheet_refs.items():
                export_data[sheet_name][column_name] = [
                    cell_ref.to_dict() for cell_ref in cell_refs
                ]
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, default=str)
        
        logger.info(f"Exported cell references to {output_path}")


def load_file_enhanced(file_path: str, use_polars: bool = True, sheet_name: Optional[str] = None) -> Tuple[Union[pd.DataFrame, pl.DataFrame], Dict[str, Any]]:
    """
    Convenience function for enhanced file loading.
    
    Args:
        file_path: Path to file to load
        use_polars: Whether to use Polars for type safety
        sheet_name: Specific sheet name for Excel files
        
    Returns:
        Tuple of (DataFrame, metadata dict)
    """
    loader = EnhancedFileLoader(use_polars=use_polars)
    return loader.load_file(file_path, sheet_name)


if __name__ == "__main__":
    # Test the enhanced loader
    import sys
    
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        df, metadata = load_file_enhanced(file_path)
        
        print(f"Loaded {metadata['total_rows']} rows, {metadata['total_columns']} columns")
        print(f"Load method: {metadata['load_method']}")
        print(f"Type inference results:")
        
        for col, inference in metadata['type_inference'].items():
            print(f"  {col}: {inference['inferred_type']} (confidence: {inference['confidence']:.2f})")
            
    else:
        print("Usage: python enhanced_loader.py <file_path>") 